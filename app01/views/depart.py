from django.shortcuts import render,redirect
from app01.models import Department
from app01.utils.pagination import Pagination
def depart_list(request):
    queryset=Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        "queryset": page_object.page_queryset,
        'page_string': page_object.html()
    }
    return render(request,'depart_list.html',context)

def depart_add(request):
    if request.method=='GET':
        return render(request,'depart_add.html')
    title=request.POST.get('title')
    Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    nid=request.GET.get("nid")
    Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

def depart_edit(request,nid):
    if request.method=="GET":
        row_object=Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html',{'row_object':row_object})
    title=request.POST.get("title")
    Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")

from openpyxl import load_workbook
def depart_multi(request):
    file_object=request.FILES.get('exc')
    wb=load_workbook(file_object)
    sheet=wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text=row[0].value
        exists=Department.objects.filter(title=text).exists()
        if not exists:
            Department.objects.create(title=text)
    return redirect('/depart/list/')